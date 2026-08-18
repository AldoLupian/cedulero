# -*- coding: utf-8 -*-
"""
Logica de extraccion y llenado de cedulas, adaptada de procesar.py para
trabajar en memoria (bytes) en vez de carpetas, para poder usarse desde
un servidor web.

Varios acuses del mismo contribuyente (mismo RFC) se juntan en una sola
cedula, con las filas ordenadas por fecha y hora de presentacion.
"""

import re
import io
import unicodedata
import difflib
from datetime import datetime
from pathlib import Path

import pdfplumber
import openpyxl

FILA_INICIAL = 39
FILA_FINAL = 63  # ultima fila disponible en la tabla de la plantilla

BANCOS_LISTA = [
    "ABACO CASA DE BOLSA, S.A. DE C.V.", "AMERICAN EXPRESS", "BANAMEX",
    "BANCO DEL ATLANTICO", "ACCIONES Y VALORES DE MEXICO, S.A. DE C.V.",
    "BANCO DEL EJERCITO, FUERZA AEREA Y ARMADA", "BANCO MERCANTIL DEL NORTE",
    "BANCO NACIONAL DE COMERCIO EXTERIOR", "BANCO NACIONAL DE OBRAS Y SERVICIOS PUBLICOS",
    "CITIBANK", "SCOTIA BANK INVERLAT", "IXE BANCO", "BANCO DEL SURESTE, S.A.",
    "BANCO DEL CREDITO RURAL DEL ITSMO", "BANCRECER", "BANREGIO", "BANCA AFIRME, S.A.",
    "AGENCIAS DEL GOBIERNO DEL ESTADO / RECAUDADORA", "BANCO DEL BAJIO, S.A.",
    "BANCO INBURSA", "HSBC", "SANTANDER SERFIN", "BBVA BANCOMER", "BANCO AZTECA",
    "BANCA MIFEL", "AGENCIAS DEL GOBIERNO DEL ESTADO / PAGO EN BANCO", "BANSI, S.A.",
    "BANCO MULTIVA", "BANCO INTERACCIONES, S.A. DE C.V.",
    "CI BANCO, S.A. INSTITUCION DE BANCA MULTIPLE",
    "BANK OF TOKYO-MITSUBISHI, UFJ (MEXICO), S.A. INST. BANCA MULTIPLE",
    "BANCO MONEX, S.A. INSTITUCION DE BANCA MULTIPLE", "BANCO BASE S.A.",
    "BANCO VE POR MAS, S.A. INST. BANCA MULTIPLE", "INTERCAM GRUPO FINANCIERO",
    "MUFG BANK MEXICO, S.A. INSTIT BANCA MULTIPLE", "PAGO EN LA TESOFE",
    "BANCO BANCREA, S.A., INSTITUCION DE BANCA MULTIPLE", "BANCO ACTINVER, S.A. DE C.V.",
    "CBM BANCO, A.A INSTITUCION DE BANCA MULTIPLE", "BANCO AUTOFIN MEXICO, S.A. DE C.V.",
    "BANCO DEL BIENESTAR, S.A.", "KPTL MEXICO BANK, S.A. INSTITUCION DE BANCA MULTIPLE",
    "BANKAOOL, S.A. INSTITUCION DE BANCA MULTIPLE",
]

UMBRAL_CONFIANZA_BANCO = 0.50

LABEL_WORDS = {
    "institucin", "de", "crdito", "fecha", "del", "pago", "lnea", "medio",
    "presentacin", "importe", "no", "pagado", "operacin", "llave", "captura",
}


def normaliza(texto):
    return re.sub(r"[^a-zA-Z]", "", texto).lower()


def normaliza_ascii(texto):
    texto = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode()
    return texto.upper()


def a_numero(texto):
    if texto is None:
        return None
    limpio = texto.replace(",", "").replace("$", "").strip()
    try:
        return int(float(limpio))
    except ValueError:
        return None


def extraer_texto_completo(pdf):
    return "\n".join(p.extract_text(x_tolerance=1, y_tolerance=1) or "" for p in pdf.pages)


MARCADOR_CONCEPTO = re.compile(r"Concepto de pago\s*\d+:[ \t]*")


def _nombre_concepto(texto_previo, bloque):
    """Arma el nombre del concepto, que puede venir partido por el marcador.

    En el formato nuevo del acuse el nombre es largo y el PDF lo parte: una
    parte queda arriba de "Concepto de pago N:" y el resto abajo.
    """
    despues = []
    for linea in bloque.strip().split("\n"):
        if ":" in linea:  # ya llegamos a "Impuesto a cargo:" y demas etiquetas
            break
        despues.append(linea.strip())
        if len(despues) >= 3:
            break

    antes = []
    for linea in reversed(texto_previo.strip().split("\n")):
        linea = linea.strip()
        if not linea or ":" in linea:
            break
        antes.insert(0, linea)
        if len(antes) >= 3:
            break

    return " ".join(p for p in antes + despues if p).strip()


def extraer_conceptos(texto_completo):
    marcadores = list(MARCADOR_CONCEPTO.finditer(texto_completo))
    conceptos = []
    for i, m in enumerate(marcadores):
        fin = marcadores[i + 1].start() if i + 1 < len(marcadores) else len(texto_completo)
        bloque = texto_completo[m.end():fin]
        bloque = re.split(r"Sello [Dd]igital|SECCI[OÓ]N|INFORMACI[OÓ]N REGISTRADA", bloque)[0]
        inicio_previo = marcadores[i - 1].end() if i else 0
        nombre = _nombre_concepto(texto_completo[inicio_previo:m.start()], bloque)
        # Anclado a inicio de linea para no confundir "A cargo" con "Cantidad a cargo".
        etiqueta = lambda nombres: re.search(
            r"^\s*(?:" + nombres + r"):\s*([\d,\.]+)", bloque, re.IGNORECASE | re.MULTILINE
        )
        m_cargo = etiqueta(r"Impuesto a cargo|A cargo")
        m_act = etiqueta(r"Actualizaciones|Parte actualizada")
        m_rec = etiqueta(r"Recargos")
        conceptos.append({
            "concepto": nombre,
            "a_cargo": a_numero(m_cargo.group(1)) if m_cargo else 0,
            "actualizaciones": a_numero(m_act.group(1)) if m_act else None,
            "recargos": a_numero(m_rec.group(1)) if m_rec else None,
        })
    return conceptos


def extraer_info_pago(pdf):
    for page in pdf.pages:
        palabras = page.extract_words(x_tolerance=1, y_tolerance=1)
        if not any("RECIBIDO" in w["text"].upper() for w in palabras):
            continue

        top_inicio = next(w["top"] for w in palabras if "RECIBIDO" in w["text"].upper())
        top_fin = next((w["top"] for w in palabras if "digital" in w["text"].lower()), 10_000)

        es_ruido_firma = lambda t: len(t) > 25 and re.match(r"^[A-Za-z0-9+/=]+$", t)

        relevantes = [
            w for w in palabras
            if top_inicio + 5 < w["top"] < top_fin - 5 and not es_ruido_firma(w["text"])
        ]

        relevantes.sort(key=lambda w: w["top"])
        filas_agrupadas = []
        for w in relevantes:
            # Tolerancia corta: en el formato nuevo del acuse la etiqueta y su
            # valor quedan a ~3.5 pt, y con un margen mayor se mezclan en una fila.
            if filas_agrupadas and abs(w["top"] - filas_agrupadas[-1][0]["top"]) <= 2.5:
                filas_agrupadas[-1].append(w)
            else:
                filas_agrupadas.append([w])

        filas_valor = []
        for palabras_fila_sin_orden in filas_agrupadas:
            palabras_fila = sorted(palabras_fila_sin_orden, key=lambda w: w["x0"])
            normalizadas = [normaliza(w["text"]) for w in palabras_fila]
            no_vacias = [n for n in normalizadas if n]
            es_etiqueta = no_vacias and all(n in LABEL_WORDS for n in no_vacias)
            if es_etiqueta:
                continue
            izquierda = " ".join(w["text"] for w in palabras_fila if w["x0"] < 300).strip()
            derecha = " ".join(w["text"] for w in palabras_fila if w["x0"] >= 300).strip()
            filas_valor.append((izquierda, derecha))

        info = {
            "banco": None, "fecha_pago": None, "linea_captura": None,
            "no_operacion": None, "llave_pago": None,
        }
        if len(filas_valor) >= 1:
            info["banco"], info["fecha_pago"] = filas_valor[0]
        if len(filas_valor) >= 2:
            info["linea_captura"], _medio = filas_valor[1]
        if len(filas_valor) >= 3:
            _importe, info["no_operacion"] = filas_valor[2]
        if len(filas_valor) >= 4:
            _vacio, info["llave_pago"] = filas_valor[3]

        return info, len(filas_valor)

    return None, 0


PALABRAS_GENERICAS_BANCO = {
    "BANCO", "BANCA", "S", "A", "SA", "DE", "C", "V", "CV", "INSTITUCION",
    "INSTIT", "INST", "MULTIPLE", "GRUPO", "FINANCIERO", "NACIONAL",
    "MEXICO", "MEX",
}


def _palabras_banco(nombre):
    limpio = re.sub(r"[^A-Z ]", " ", normaliza_ascii(nombre))
    return [w for w in limpio.split() if w not in PALABRAS_GENERICAS_BANCO]


def _normaliza_banco(nombre):
    return "".join(_palabras_banco(nombre))


def _score_banco(objetivo, opcion_norm, palabras_objetivo, palabras_opcion):
    """Pesa sobre todo las palabras distintivas compartidas.

    El parecido de subcadena solo desempata: por si solo premia a los nombres
    cortos de la lista (p. ej. "BBVA Mexico" se parecia mas a BANAMEX que a
    BBVA BANCOMER por el "AMEX" comun).
    """
    comunes = sum(len(p) for p in set(palabras_objetivo) & set(palabras_opcion))
    letras_opcion = sum(len(p) for p in palabras_opcion) or 1
    letras_objetivo = sum(len(p) for p in palabras_objetivo) or 1
    # Promedio de que tanto se cubren mutuamente: "BBVA" cubre todo el nombre
    # del PDF aunque solo sea la mitad de "BBVA BANCOMER".
    score_palabras = (comunes / letras_opcion + comunes / letras_objetivo) / 2

    sm = difflib.SequenceMatcher(None, objetivo, opcion_norm)
    match = sm.find_longest_match(0, len(objetivo), 0, len(opcion_norm))
    score_subcadena = match.size / max(len(opcion_norm), len(objetivo), 1)

    return max(0.7 * score_palabras + 0.3 * score_subcadena, score_subcadena)


def mapear_banco(nombre_pdf):
    if not nombre_pdf:
        return None, 0.0
    palabras_objetivo = _palabras_banco(nombre_pdf)
    objetivo = "".join(palabras_objetivo)
    mejor_opcion, mejor_score = None, 0.0
    for opcion in BANCOS_LISTA:
        palabras_opcion = _palabras_banco(opcion)
        opcion_norm = "".join(palabras_opcion)
        if not opcion_norm:
            continue
        score = _score_banco(objetivo, opcion_norm, palabras_objetivo, palabras_opcion)
        if score > mejor_score:
            mejor_opcion, mejor_score = opcion, score
    return mejor_opcion, mejor_score


def mapear_tipo_impuesto(concepto):
    c = normaliza_ascii(concepto)
    retencion = "RETEN" in c
    if "IEPS" in c:
        return "IEPS"
    if "IVA" in c or "VALOR AGREGADO" in c:
        return "IVA RET" if retencion else "IVA"
    if "ISR" in c or "SOBRE LA RENTA" in c:
        return "ISR RET" if retencion else "ISR"
    return "OTR RET" if retencion else "OTR"


def parsear_fecha(texto):
    if not texto:
        return None
    m = re.search(r"(\d{2})/(\d{2})/(\d{4})", texto)
    if not m:
        return None
    dia, mes, anio = m.groups()
    return datetime(int(anio), int(mes), int(dia))


def extraer_datos_contribuyente(texto_completo):
    m_rfc = re.search(r"RFC:\s*(\S+)", texto_completo)
    m_razon = re.search(
        r"Denominaci[oó]n o raz[oó]n social:\s*(.+)",
        texto_completo, re.IGNORECASE,
    )
    rfc = m_rfc.group(1).strip() if m_rfc else None
    razon_social = m_razon.group(1).strip() if m_razon else None
    return rfc, razon_social


MESES = {
    "enero": "01", "febrero": "02", "marzo": "03", "abril": "04",
    "mayo": "05", "junio": "06", "julio": "07", "agosto": "08",
    "septiembre": "09", "setiembre": "09", "octubre": "10",
    "noviembre": "11", "diciembre": "12",
}


def extraer_periodo_declaracion(texto_completo):
    m_periodo = re.search(
        r"Per[ií]odo de la declaraci[oó]n:\s*(\S+)",
        texto_completo, re.IGNORECASE,
    )
    m_ejercicio = re.search(r"Ejercicio:\s*(\d{4})", texto_completo, re.IGNORECASE)
    if not m_periodo or not m_ejercicio:
        return None
    mes = MESES.get(m_periodo.group(1).strip().lower())
    if not mes:
        return None
    anio = m_ejercicio.group(1)[-2:]
    return f"{mes}-{anio}"


def extraer_fecha_presentacion(texto_completo):
    """Fecha y hora en que se presento la declaracion: define el orden de las filas."""
    m = re.search(
        r"Fecha y hora de presentaci[oó]n:?\s*(\d{2})/(\d{2})/(\d{4})\s*-?\s*(\d{1,2}):(\d{2})",
        texto_completo, re.IGNORECASE,
    )
    if not m:
        return None
    dia, mes, anio, hora, minuto = m.groups()
    return datetime(int(anio), int(mes), int(dia), int(hora), int(minuto))


def extraer_datos_pdf(nombre_archivo, pdf_bytes):
    """Lee un acuse y devuelve sus datos normalizados, sin tocar la plantilla.

    Cuando el PDF no se puede leer devuelve el motivo en 'error'; el llamador
    decide si lo agrupa con otros acuses o lo reporta como error.
    """
    avisos = []

    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            texto_completo = extraer_texto_completo(pdf)
            conceptos = extraer_conceptos(texto_completo)
            info_pago, _n_filas_valor = extraer_info_pago(pdf)
            periodo_declaracion = extraer_periodo_declaracion(texto_completo)
            rfc, razon_social = extraer_datos_contribuyente(texto_completo)
            fecha_presentacion = extraer_fecha_presentacion(texto_completo)
    except Exception as e:
        return {"nombre": nombre_archivo, "error": f"No se pudo abrir el PDF: {e}",
                "avisos": [], "conceptos": []}

    if not conceptos:
        return {"nombre": nombre_archivo,
                "error": "No se encontraron conceptos de pago en el PDF. "
                         "Verifica que sea un acuse de pago con linea de captura del SAT.",
                "avisos": [], "conceptos": []}

    if info_pago is None:
        avisos.append("No se encontro la seccion 'Informacion del pago recibido'. "
                      "No se pudieron llenar banco, fecha, no. de operacion ni llave de pago.")
        info_pago = {"banco": None, "fecha_pago": None, "linea_captura": None,
                     "no_operacion": None, "llave_pago": None}

    if fecha_presentacion is None:
        avisos.append("No se encontro la fecha y hora de presentacion; este acuse quedo "
                      "al final de la cedula.")

    banco_mapeado, score_banco = mapear_banco(info_pago["banco"])
    if info_pago["banco"] and score_banco < UMBRAL_CONFIANZA_BANCO:
        avisos.append("No se pudo identificar el banco con certeza "
                      f"(\"{info_pago['banco']}\"). Se dejo el texto tal cual del PDF; "
                      "revisa y selecciona manualmente en la lista del Excel.")
        banco_final = info_pago["banco"]
    else:
        banco_final = banco_mapeado

    # Solo las lineas con importe; las que quedan en cero no se capturan.
    filas = []
    for c in conceptos:
        if (c["a_cargo"] or 0) + (c["actualizaciones"] or 0) + (c["recargos"] or 0) == 0:
            continue
        filas.append({
            "tipo_impuesto": mapear_tipo_impuesto(c["concepto"]),
            "a_cargo": c["a_cargo"],
            "actualizaciones": c["actualizaciones"],
            "recargos": c["recargos"],
        })

    return {
        "nombre": nombre_archivo,
        "error": None,
        "avisos": avisos,
        "rfc": rfc,
        "razon_social": razon_social,
        "fecha_presentacion": fecha_presentacion,
        "periodo_pago": periodo_declaracion,
        "banco": banco_final,
        "fecha_pago": parsear_fecha(info_pago["fecha_pago"]),
        "no_operacion": a_numero(info_pago["no_operacion"]) or info_pago["no_operacion"],
        "llave_pago": info_pago["llave_pago"],
        "linea_captura": (info_pago["linea_captura"] or "").replace(" ", ""),
        "conceptos": filas,
    }


def clave_contribuyente(datos):
    """Los acuses se agrupan por RFC; sin RFC cada acuse va en su propia cedula."""
    rfc = (datos.get("rfc") or "").strip().upper()
    return rfc or "__sin_rfc__::" + datos["nombre"]


def agrupar_por_contribuyente(lista_datos):
    """Agrupa por RFC conservando el orden en que llegaron los acuses."""
    grupos = {}
    for datos in lista_datos:
        grupos.setdefault(clave_contribuyente(datos), []).append(datos)
    return list(grupos.values())


def ordenar_por_presentacion(acuses):
    """Ordena por fecha y hora de presentacion; los que no la traen van al final."""
    return sorted(
        acuses,
        key=lambda d: (d["fecha_presentacion"] is None, d["fecha_presentacion"] or datetime.max),
    )


def nombre_cedula(acuses):
    """Nombre visible de la cedula: el del acuse mas antiguo del contribuyente."""
    base = Path(acuses[0]["nombre"]).stem
    return base if len(acuses) == 1 else f"{base} (+{len(acuses) - 1})"


def generar_cedula(acuses, plantilla_path):
    """Llena una sola cedula con todos los acuses de un mismo contribuyente."""
    acuses = ordenar_por_presentacion(acuses)
    varios = len(acuses) > 1

    avisos = []
    for a in acuses:
        prefijo = f"{a['nombre']}: " if varios else ""
        avisos.extend(prefijo + aviso for aviso in a["avisos"])

    if len({a["razon_social"] for a in acuses if a["razon_social"]}) > 1:
        avisos.append("Los acuses traen razones sociales distintas aunque comparten RFC. "
                      "Se uso la del acuse mas antiguo; revisa el encabezado.")

    filas_totales = sum(len(a["conceptos"]) for a in acuses)
    disponibles = FILA_FINAL - FILA_INICIAL + 1
    if filas_totales > disponibles:
        avisos.append(f"Los acuses traen {filas_totales} conceptos pero la plantilla solo "
                      f"tiene {disponibles} filas disponibles. Se recorto.")

    wb = openpyxl.load_workbook(plantilla_path)
    ws = wb.active

    ws["T4"] = acuses[0]["rfc"]
    ws["AN6"] = acuses[0]["razon_social"]

    importe_total = 0
    fila = FILA_INICIAL
    for acuse in acuses:
        for c in acuse["conceptos"]:
            if fila > FILA_FINAL:
                break
            ws[f"C{fila}"] = "C"
            ws[f"D{fila}"] = "AUTOCORRECCION TOTAL"
            ws[f"K{fila}"] = "P"
            ws[f"M{fila}"] = acuse["no_operacion"]
            ws[f"S{fila}"] = acuse["llave_pago"]
            ws[f"Y{fila}"] = acuse["linea_captura"]
            ws[f"AI{fila}"] = acuse["banco"]
            ws[f"AN{fila}"] = c["tipo_impuesto"]
            ws[f"AQ{fila}"] = acuse["periodo_pago"]
            ws[f"AT{fila}"] = acuse["fecha_pago"]
            ws[f"AX{fila}"] = c["a_cargo"]
            importe_total += c["a_cargo"] or 0
            if c["actualizaciones"]:
                ws[f"BB{fila}"] = c["actualizaciones"]
                importe_total += c["actualizaciones"]
            if c["recargos"]:
                ws[f"BL{fila}"] = c["recargos"]
                importe_total += c["recargos"]
            fila += 1

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    bancos = {a["banco"] for a in acuses if a["banco"]}
    fechas = {a["fecha_pago"] for a in acuses if a["fecha_pago"]}
    fecha_pago = min(fechas) if fechas else None

    return {
        "nombre": nombre_cedula(acuses),
        "rfc": acuses[0]["rfc"],
        "razon_social": acuses[0]["razon_social"],
        "archivos": [a["nombre"] for a in acuses],
        "estado": "advertencia" if avisos else "correcto",
        "avisos": avisos,
        "banco": bancos.pop() if len(bancos) == 1 else ("Varios" if bancos else None),
        "fecha_pago": fecha_pago.strftime("%m-%y") if fecha_pago else None,
        "importe_total": importe_total,
        "filas": min(filas_totales, disponibles),
        "xlsx_bytes": buffer.getvalue(),
    }


def procesar_pdfs(archivos, plantilla_path):
    """archivos: lista de (nombre, bytes). Devuelve (cedulas, errores).

    Los acuses del mismo RFC terminan en una sola cedula, con las filas
    ordenadas por fecha y hora de presentacion.
    """
    leidos, errores = [], []
    for nombre, contenido in archivos:
        datos = extraer_datos_pdf(nombre, contenido)
        if datos["error"]:
            errores.append({"nombre": nombre, "estado": "error", "avisos": [datos["error"]]})
        elif not datos["conceptos"]:
            errores.append({"nombre": nombre, "estado": "error",
                            "avisos": ["Todos los conceptos de pago de este acuse estan en cero."]})
        else:
            leidos.append(datos)

    cedulas = [generar_cedula(grupo, plantilla_path)
               for grupo in agrupar_por_contribuyente(leidos)]
    return cedulas, errores
